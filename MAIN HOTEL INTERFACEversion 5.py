from tkinter import *
from tkinter import ttk
from tkinter import filedialog, messagebox
import random
import time
import tkinter

from openpyxl import Workbook
import os



#................................................import requests...................................................................
root = Tk()
root.geometry('1530x800+0+0')
root.resizable(0, 0)
root.title('Restaurant Management System ')
root.config(bg='firebrick4')
topFrame = Frame(root, bd=10, relief=RIDGE, bg='firebrick4')
topFrame.pack(side=TOP)
labelTitle = Label(topFrame, text='HOTEL MANAGEMENT SYSTEM', font=('arial', 30, 'bold'), fg='yellow', bd=9,bg='red4', width=51)
labelTitle.grid(row=0, column=0)


#..................................................... Reset Function....................................................................

def reset():
    textReceipt.delete(1.0, END)
    e_daal.set('0')
    e_roti.set('0')
    e_sabji.set('0')
    e_fish.set('0')
    e_kebab.set('0')
    e_chawal.set('0')
    e_mutton.set('0')
    e_paneer.set('0')
    e_chicken.set('0')

    e_nonacs.set('0')
    e_acs.set('0')
    e_nonacd.set('0')
    e_acd.set('0')
    e_tire1.set('0')
    e_table1.set('0')
    e_table2.set('0')
    e_table3.set('0')
    e_table4.set('0')

    textroti.config(state=DISABLED)
    textdaal.config(state=DISABLED)
    textsabji.config(state=DISABLED)
    textfish.config(state=DISABLED)
    textkebab.config(state=DISABLED)
    textpaneer.config(state=DISABLED)
    textchicken.config(state=DISABLED)
    textmutton.config(state=DISABLED)
    textchawal.config(state=DISABLED)

    textnonacs.config(state=DISABLED)
    textacs.config(state=DISABLED)
    textnonacd.config(state=DISABLED)
    textacd.config(state=DISABLED)
    texttire1.config(state=DISABLED)
    texttable1.config(state=DISABLED)
    texttable2.config(state=DISABLED)
    texttable3.config(state=DISABLED)
    texttable4.config(state=DISABLED)

    var1.set(0)
    var2.set(0)
    var3.set(0)
    var4.set(0)
    var5.set(0)
    var6.set(0)
    var7.set(0)
    var8.set(0)
    var9.set(0)
    var10.set(0)
    var11.set(0)
    var12.set(0)
    var13.set(0)
    var14.set(0)
    var15.set(0)
    var16.set(0)
    var17.set(0)
    var18.set(0)
    
    costofroomvar.set('')
    costoffoodvar.set('')
    subtotalvar.set('')
    servicetaxvar.set('')
    totalcostvar.set('')
    lbl_name.delete(0, 'end') 
    lbl_add.delete(0,'end')
    lbl_no.delete(0,'end')
    lbl_email.delete(0,'end')
    lbl_id.delete(0,'end')

#..............................................save fumction......................................................................
def save():
    if textReceipt.get(1.0, END) == '\n':
        pass
    else:
        url = filedialog.asksaveasfile(mode='w', defaultextension='.csv')
        if url == None:
            pass
        else:

            bill_data = textReceipt.get(1.0, END)
            url.write(bill_data)
            url.close()
            messagebox.showinfo('Information', 'Your Bill Is Succesfully Saved')

#................................................total cost function.................................................................
def totalcost():
    global priceofFood, priceofroom, subtotalofItems
    if var1.get() != 0 or var2.get() != 0 or var3.get() != 0 or var4.get() != 0 or var5.get() != 0 or \
            var6.get() != 0 or var7.get() != 0 or var8.get() != 0 or var9.get() != 0 or var10.get() != 0 or \
            var11.get() != 0 or var12.get() != 0 or var13.get() != 0 or var14.get() != 0 or var15.get() != 0 or \
            var16.get() != 0 or var17.get() != 0 or var18.get() != 0 :

        item1 = int(e_roti.get())
        item2 = int(e_daal.get())
        item3 = int(e_fish.get())
        item4 = int(e_sabji.get())
        item5 = int(e_kebab.get())
        item6 = int(e_chawal.get())
        item7 = int(e_mutton.get())
        item8 = int(e_paneer.get())
        item9 = int(e_chicken.get())

        item10 = int(e_nonacs.get())
        item11 = int(e_acs.get())
        item12 = int(e_nonacd.get())
        item13 = int(e_acd.get())
        item14 = int(e_tire1.get())
        item15 = int(e_table1.get())
        item16 = int(e_table2.get())
        item17 = int(e_table3.get())
        item18 = int(e_table4.get())
        
        
        
        priceofFood = (item1 * 10) + (item2 * 60) + (item3 * 100) + (item4 * 50) + (item5 * 40) + (item6 * 30) + (
                    item7 * 120) \
                      + (item8 * 100) + (item9 * 120)

        priceofroom = (item10 * 1999) + (item11 * 2499) + (item12 *2999 ) + (item13 * 3499) + (item14 * 4999) + (item15 * 199) \
                        + (item16 * 299) + (item17 * 399) + (item18 * 499)

        
        costoffoodvar.set(str(priceofFood) + ' Rs')
        costofroomvar.set(str(priceofroom) + ' Rs')
        

        subtotalofItems = priceofFood + priceofroom
        subtotalvar.set(str(subtotalofItems) + ' Rs')

        servicetaxvar.set('50 Rs')

        tottalcost = subtotalofItems + 50
        totalcostvar.set(str(tottalcost) + ' Rs')

    else:
        messagebox.showerror('Error', 'No Item Is selected')


def roti():
    if var1.get() == 1:
        textroti.config(state=NORMAL)
        textroti.delete(0, END)
        textroti.focus()
    else:
        textroti.config(state=DISABLED)
        e_roti.set('0')


def daal():
    if var2.get() == 1:
        textdaal.config(state=NORMAL)
        textdaal.delete(0, END)
        textdaal.focus()

    else:
        textdaal.config(state=DISABLED)
        e_daal.set('0')


def fish():
    if var3.get() == 1:
        textfish.config(state=NORMAL)
        textfish.delete(0, END)
        textfish.focus()

    else:
        textfish.config(state=DISABLED)
        e_fish.set('0')


def sabji():
    if var4.get() == 1:
        textsabji.config(state=NORMAL)
        textsabji.focus()
        textsabji.delete(0, END)
    elif var4.get() == 0:
        textsabji.config(state=DISABLED)
        e_sabji.set('0')


def kebab():
    if var5.get() == 1:
        textkebab.config(state=NORMAL)
        textkebab.focus()
        textkebab.delete(0, END)
    elif var5.get() == 0:
        textkebab.config(state=DISABLED)
        e_kebab.set('0')


def chawal():
    if var6.get() == 1:
        textchawal.config(state=NORMAL)
        textchawal.focus()
        textchawal.delete(0, END)
    elif var6.get() == 0:
        textchawal.config(state=DISABLED)
        e_chawal.set('0')


def mutton():
    if var7.get() == 1:
        textmutton.config(state=NORMAL)
        textmutton.focus()
        textmutton.delete(0, END)
    elif var7.get() == 0:
        textmutton.config(state=DISABLED)
        e_mutton.set('0')


def paneer():
    if var8.get() == 1:
        textpaneer.config(state=NORMAL)
        textpaneer.focus()
        textpaneer.delete(0, END)
    elif var8.get() == 0:
        textpaneer.config(state=DISABLED)
        e_paneer.set('0')


def chicken():
    if var9.get() == 1:
        textchicken.config(state=NORMAL)
        textchicken.focus()
        textchicken.delete(0, END)
    elif var9.get() == 0:
        textchicken.config(state=DISABLED)
        e_chicken.set('0')


def nonacs():
    if var10.get() == 1:
        textnonacs.config(state=NORMAL)
        textnonacs.focus()
        textnonacs.delete(0, END)
    elif var10.get() == 0:
        textnonacs.config(state=DISABLED)
        e_nonacs.set('0')


def acs():
    if var11.get() == 1:
        textacs.config(state=NORMAL)
        textacs.focus()
        textacs.delete(0, END)
    elif var11.get() == 0:
        textacs.config(state=DISABLED)
        e_acs.set('0')


def nonacd():
    if var12.get() == 1:
        textnonacd.config(state=NORMAL)
        textnonacd.focus()
        textnonacd.delete(0, END)
    elif var12.get() == 0:
        textnonacd.config(state=DISABLED)
        e_nonacd.set('0')

def acd():
    if var13.get() == 1:
        textacd.config(state=NORMAL)
        textacd.focus()
        textacd.delete(0, END)
    elif var13.get() == 0:
        textacd.config(state=DISABLED)
        e_acd.set('0')


def tire1():
    if var14.get() == 1:
        texttire1.config(state=NORMAL)
        texttire1.focus()
        texttire1.delete(0, END)
    elif var14.get() == 0:
        texttire1.config(state=DISABLED)
        e_tire1.set('0')


def table1():
    if var15.get() == 1:
        texttable1.config(state=NORMAL)
        texttable1.focus()
        texttable1.delete(0, END)
    elif var15.get() == 0:
        texttable1.config(state=DISABLED)
        e_table1.set('0')



def table2():
    if var16.get() == 1:
        texttable2.config(state=NORMAL)
        texttable2.focus()
        texttable2.delete(0, END)
    elif var16.get() == 0:
        texttable2.config(state=DISABLED)
        e_table2.set('0')


def table3():
    if var17.get() == 1:
        texttable3.config(state=NORMAL)
        texttable3.focus()
        texttable3.delete(0, END)
    elif var17.get() == 0:
        texttable3.config(state=DISABLED)
        e_table3.set('0')


def table4():
    if var18.get() == 1:
        texttable4.config(state=NORMAL)
        texttable4.focus()
        texttable4.delete(0, END)
    elif var18.get() == 0:
        texttable4.config(state=DISABLED)
        e_table4.set('0')



#.................................................Receipt function................................................................
def receipt():
    global billnumber, date
    if costoffoodvar.get() != '' or costofcakesvar.get() != '' or costofroomvar.get() != '':
        textReceipt.delete(1.0, END)
        x = random.randint(100, 10000)

        date = time.strftime('%d/%m/%Y')
        textReceipt.insert(END, 'Receipt Ref:\t\t' + f'ID NO:\t{(e_entryno.get())}' + '\t' + date + '\n')

        textReceipt.insert(END, f'Name:\t{(e_name.get())}\n')
        textReceipt.insert(END, f'Mob NO:\t{(e_mobno.get())}\n\n')
        textReceipt.insert(END, '***************************************************************\n')
        textReceipt.insert(END, 'Items:\t\t Cost Of Items(Rs)\n')
        textReceipt.insert(END, '***************************************************************\n')
        if e_roti.get() != '0':
            textReceipt.insert(END, f'Roti\t\t\t{int(e_roti.get()) * 10}\n\n')

        if e_daal.get() != '0':
            textReceipt.insert(END, f'Daal\t\t\t{int(e_daal.get()) * 60}\n\n')

        if e_fish.get() != '0':
            textReceipt.insert(END, f'Fish\t\t\t{int(e_fish.get()) * 100}\n\n')

        if e_chawal.get() != '0':
            textReceipt.insert(END, f'Chawal:\t\t\t{int(e_chawal.get()) * 30}\n\n')

        if e_sabji.get() != '0':
            textReceipt.insert(END, f'Sabji:\t\t\t{int(e_sabji.get()) * 50}\n\n')

        if e_paneer.get() != '0':
            textReceipt.insert(END, f'Paneer:\t\t\t{int(e_paneer.get()) * 100}\n\n')

        if e_kebab.get() != '0':
            textReceipt.insert(END, f'Kebab:\t\t\t{int(e_kebab.get()) * 40}\n\n')

        if e_chicken.get() != '0':
            textReceipt.insert(END, f'Chicken:\t\t\t{int(e_chicken.get()) * 120}\n\n')

        if e_mutton.get() != '0':
            textReceipt.insert(END, f'Mutton:\t\t\t{int(e_mutton.get()) * 120}\n\n')

        if e_nonacs.get() != '0':
            textReceipt.insert(END, f'NON-AC Single:\t\t\t{int(e_nonacs.get()) * 1999}\n\n')

        if e_acs.get() != '0':
            textReceipt.insert(END, f'AC Single:\t\t\t{int(e_acs.get()) * 2499}\n\n')

        if e_nonacd.get() != '0':
            textReceipt.insert(END, f'NON-AC Double:\t\t\t{int(e_nonacd.get()) * 2999}\n\n')

        if e_acd.get() != '0':
            textReceipt.insert(END, f'AC Double:\t\t\t{int(e_acd.get()) * 3499}\n\n')

        if e_tire1.get() != '0':
            textReceipt.insert(END, f'Tire1:\t\t\t{int(e_tire1.get()) * 4999}\n\n')

        if e_table1.get() != '0':
            textReceipt.insert(END, f'Table1:\t\t\t{int(e_table1.get()) * 199}\n\n')


        if e_table2.get() != '0':
            textReceipt.insert(END, f'Table2:\t\t\t{int(e_table2.get()) * 299}\n\n')

        if e_table3.get() != '0':
            textReceipt.insert(END, f'Table3:\t\t\t{int(e_table3.get()) * 399}\n\n')

        if e_table4.get() != '0':
            textReceipt.insert(END, f'Table4:\t\t\t{int(e_table4.get()) * 499}\n\n')

        
        textReceipt.insert(END, '***************************************************************\n')
        if costoffoodvar.get() != '0 Rs':
            textReceipt.insert(END, f'Cost Of Food\t\t\t{priceofFood}Rs\n\n')
        if costofroomvar.get() != '0 Rs':
            textReceipt.insert(END, f'Cost Of Room\t\t\t{priceofroom}Rs\n\n')
        

        textReceipt.insert(END, f'Sub Total\t\t\t{subtotalofItems}Rs\n\n')
        textReceipt.insert(END, f'Service Tax\t\t\t{50}Rs\n\n')
        textReceipt.insert(END, f'Total Cost\t\t\t{subtotalofItems + 50}Rs\n\n')
        textReceipt.insert(END, '***************************************************************\n')

    else:
        messagebox.showerror('Error', 'No Item Is selected')


#....................................................... frames.....................................................................

menuFrame = Frame(root, bd=10, relief=RIDGE, bg='firebrick4')
menuFrame.pack(side=LEFT)

costFrame = Frame(menuFrame, bd=4, relief=RIDGE, bg='firebrick4', pady=10)
costFrame.pack(side=BOTTOM)

foodFrame = LabelFrame(menuFrame, text='FOOD', font=('arial', 19, 'bold'), bd=10, relief=RIDGE, fg='red4', )
foodFrame.pack(side=LEFT)

roomFrame = LabelFrame(menuFrame, text='ROOM', font=('arial', 19, 'bold'), bd=10, relief=RIDGE, fg='red4')
roomFrame.pack(side=LEFT)

rightFrame = Frame(root, bd=12, relief=RIDGE, bg='red4')
rightFrame.place(x=870,y=140,width=650,height=600)

recieptFrame = Frame(rightFrame, bd=10, relief=RIDGE, bg='red4')
recieptFrame.place(x=20,y=250,width=510,height=330)

buttonFrame = Frame(rightFrame, bd=5, relief=RIDGE, bg='red4')
buttonFrame.place(x=20,y=530,width=505)



# .............................................................Variables..............................................................

var1 = IntVar()
var2 = IntVar()
var3 = IntVar()
var4 = IntVar()
var5 = IntVar()
var6 = IntVar()
var7 = IntVar()
var8 = IntVar()
var9 = IntVar()
var10 = IntVar()
var11 = IntVar()
var12 = IntVar()
var13 = IntVar()
var14 = IntVar()
var15 = IntVar()
var16 = IntVar()
var17 = IntVar()
var18 = IntVar()


e_roti = StringVar()
e_daal = StringVar()
e_sabji = StringVar()
e_chawal = StringVar()
e_fish = StringVar()
e_mutton = StringVar()
e_kebab = StringVar()
e_chicken = StringVar()
e_paneer = StringVar()

e_nonacs = StringVar()
e_acs = StringVar()
e_nonacd = StringVar()
e_acd = StringVar()
e_tire1 = StringVar()
e_table1 = StringVar()
e_table2 = StringVar()
e_table3 = StringVar()
e_table4 = StringVar()
e_name = StringVar()
e_entryno = StringVar()
e_mobno = StringVar()


costoffoodvar = StringVar()
costofroomvar = StringVar()
costofcakesvar = StringVar()
subtotalvar = StringVar()
servicetaxvar = StringVar()
totalcostvar = StringVar()

e_roti.set('0')
e_daal.set('0')
e_sabji.set('0')
e_fish.set('0')
e_kebab.set('0')
e_chawal.set('0')
e_mutton.set('0')
e_chicken.set('0')
e_paneer.set('0')

e_nonacs.set('0')
e_acs.set('0')
e_nonacd.set('0')
e_acd.set('0')
e_tire1.set('0')
e_table1.set('0')
e_table2.set('0')
e_table3.set('0')
e_table4.set('0')



#..............................................................FOOD.................................................................

roti = Checkbutton(foodFrame, text='Roti          [RS 10]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var1
                   , command=roti)
roti.grid(row=0, column=0, sticky=W)

daal = Checkbutton(foodFrame, text='Daal          [RS 60]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var2
                   , command=daal)
daal.grid(row=1, column=0, sticky=W)

fish = Checkbutton(foodFrame, text='Fish          [RS 100]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var3
                   , command=fish)
fish.grid(row=2, column=0, sticky=W)

sabji = Checkbutton(foodFrame, text='Sabji         [RS 50]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var4
                    , command=sabji)
sabji.grid(row=3, column=0, sticky=W)

kebab = Checkbutton(foodFrame, text='kebab        [RS 40]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var5
                    , command=kebab)
kebab.grid(row=4, column=0, sticky=W)

chawal = Checkbutton(foodFrame, text='Chawal      [RS 30]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var6
                     , command=chawal)
chawal.grid(row=5, column=0, sticky=W)

mutton = Checkbutton(foodFrame, text='Mutton      [RS 120]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var7,
                     command=mutton)
mutton.grid(row=6, column=0, sticky=W)

paneer = Checkbutton(foodFrame, text='Paneer      [RS 100]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var8
                     , command=paneer)
paneer.grid(row=7, column=0, sticky=W)

chicken = Checkbutton(foodFrame, text='Chicken    [RS 120 ]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var9
                      , command=chicken)
chicken.grid(row=8, column=0, sticky=W)

#....................................................... Entry Fields for Food Items.......................................................

textroti = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_roti)
textroti.grid(row=0, column=1)

textdaal = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_daal)
textdaal.grid(row=1, column=1)

textfish = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_fish)
textfish.grid(row=2, column=1)

textsabji = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_sabji)
textsabji.grid(row=3, column=1)

textkebab = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_kebab)
textkebab.grid(row=4, column=1)

textchawal = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_chawal)
textchawal.grid(row=5, column=1)

textmutton = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_mutton)
textmutton.grid(row=6, column=1)

textpaneer = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_paneer)
textpaneer.grid(row=7, column=1)

textchicken = Entry(foodFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_chicken)
textchicken.grid(row=8, column=1)

#........................................................... rooms....................................................................

nonacs = Checkbutton(roomFrame, text='NON-AC Single  [RS 1999]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var10
                    , command=nonacs)
nonacs.grid(row=0, column=0, sticky=W)

acs = Checkbutton(roomFrame, text='AC Single           [RS 2499]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var11
                     , command=acs)
acs.grid(row=1, column=0, sticky=W)

nonacd = Checkbutton(roomFrame, text='NON-AC Double  [RS 2999]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var12
                     , command=nonacd)
nonacd.grid(row=2, column=0, sticky=W)

acd = Checkbutton(roomFrame, text='AC Double           [RS 3499]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var13
                       , command=acd)
acd.grid(row=3, column=0, sticky=W)

tire1 = Checkbutton(roomFrame, text='Tire1                    [RS 4999]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var14
                       , command=tire1)
tire1.grid(row=4, column=0, sticky=W)

table1 = Checkbutton(roomFrame, text='Table1                 [RS 199]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0, variable=var15
                       , command=table1)
table1.grid(row=5, column=0, sticky=W)

table2 = Checkbutton(roomFrame, text='Table2                 [RS 299]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0,
                        variable=var16
                        , command=table2)
table2.grid(row=6, column=0, sticky=W)

table3 = Checkbutton(roomFrame, text='Table3                 [RS 399]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0,
                        variable=var17
                        , command=table3)
table3.grid(row=7, column=0, sticky=W)

table4 = Checkbutton(roomFrame, text='Table4                  [RS 499]', font=('arial', 18, 'bold'), onvalue=1, offvalue=0,
                         variable=var18
                         , command=table4)
table4.grid(row=8, column=0, sticky=W)

#..................................................... entry fields for room..............................................................

textnonacs = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_nonacs)
textnonacs.grid(row=0, column=1)

textacs = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_acs)
textacs.grid(row=1, column=1)

textnonacd = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_nonacd)
textnonacd.grid(row=2, column=1)

textacd = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_acd)
textacd.grid(row=3, column=1)


texttire1 = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_tire1)
texttire1.grid(row=4, column=1)

texttable1 = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_table1)
texttable1.grid(row=5, column=1)


texttable2 = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_table2)
texttable2.grid(row=6, column=1)

texttable3 = Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_table3)
texttable3.grid(row=7, column=1)

texttable4= Entry(roomFrame, font=('arial', 18, 'bold'), bd=7, width=6, state=DISABLED, textvariable=e_table4)
texttable4.grid(row=8, column=1)



#............................................ costlabels & entry fields..................................................................

labelCostofFood = Label(costFrame, text='Cost of Food', font=('arial', 16, 'bold'), bg='firebrick4', fg='white')
labelCostofFood.grid(row=0, column=0)

textCostofFood = Entry(costFrame, font=('arial', 16, 'bold'), bd=6, width=14, state='readonly',
                       textvariable=costoffoodvar)
textCostofFood.grid(row=0, column=1, padx=41)

labelCostofroom = Label(costFrame, text='Cost of room', font=('arial', 16, 'bold'), bg='firebrick4', fg='white')
labelCostofroom.grid(row=1, column=0)

textCostofroom = Entry(costFrame, font=('arial', 16, 'bold'), bd=6, width=14, state='readonly',
                         textvariable=costofroomvar)
textCostofroom.grid(row=1, column=1, padx=41)


labelSubTotal = Label(costFrame, text='Sub Total', font=('arial', 16, 'bold'), bg='firebrick4', fg='white')
labelSubTotal.grid(row=0, column=2)

textSubTotal = Entry(costFrame, font=('arial', 16, 'bold'), bd=6, width=14, state='readonly', textvariable=subtotalvar)
textSubTotal.grid(row=0, column=3, padx=41)

labelServiceTax = Label(costFrame, text='Service Tax', font=('arial', 16, 'bold'), bg='firebrick4', fg='white')
labelServiceTax.grid(row=1, column=2)

textServiceTax = Entry(costFrame, font=('arial', 16, 'bold'), bd=6, width=14, state='readonly',
                       textvariable=servicetaxvar)
textServiceTax.grid(row=1, column=3, padx=41)

labelTotalCost = Label(costFrame, text='Total Cost', font=('arial', 16, 'bold'), bg='firebrick4', fg='white')
labelTotalCost.grid(row=2, column=2)

textTotalCost = Entry(costFrame, font=('arial', 16, 'bold'), bd=6, width=14, state='readonly',
                      textvariable=totalcostvar)
textTotalCost.grid(row=2, column=3, padx=41)

#..................................................... Buttons.........................................................................

buttonTotal = Button(buttonFrame, text='Total', font=('arial', 14, 'bold'), fg='white', bg='red4', bd=3, padx=5,
                     command=totalcost)
buttonTotal.grid(row=0, column=0)

buttonReceipt = Button(buttonFrame, text='Receipt', font=('arial', 14, 'bold'), fg='white', bg='red4', bd=3, padx=5
                       , command=receipt)
buttonReceipt.grid(row=0, column=1)

buttonSave = Button(buttonFrame, text='Save', font=('arial', 14, 'bold'), fg='white', bg='red4', bd=3, padx=5
                    , command=save)
buttonSave.grid(row=0, column=2)

buttonReset = Button(buttonFrame, text='Reset', font=('arial', 14, 'bold'), fg='white', bg='red4', bd=3, padx=5,
                     command=reset)
buttonReset.grid(row=0, column=3)

# textarea for receipt

textReceipt = Text(recieptFrame, font=('arial', 12, 'bold'), bd=3, width=42, height=14)
textReceipt.grid(row=0, column=0)



#.............................................customer details...............................................................



lbl_name=Label(rightFrame,text=" Name:",font=("times new roman",20,"bold"),fg="white",bg="red4").place(x=30,y=10)
lbl_name=Entry(rightFrame,font=("times new roman",15),bg="white",textvariable=e_name)
lbl_name.place(x=200,y=10,width=400,height=35)
   
lbl_add=Label(rightFrame,text="Address:",font=("times new roman",20,"bold"),fg="white",bg="red4").place(x=30,y=60)
lbl_add=Entry(rightFrame,font=("times new roman",15),bg="white")
lbl_add.place(x=200,y=60,width=400,height=35)
   

lbl_no=Label(rightFrame,text="Mob NO:",font=("times new roman",20,"bold"),fg="white",bg="red4").place(x=30,y=110)
lbl_no=Entry(rightFrame,font=("times new roman",15),bg="white",textvariable=e_mobno)
lbl_no.place(x=200,y=110,width=400,height=35)


lbl_email=Label(rightFrame,text="Email ID:",font=("times new roman",20,"bold"),fg="white",bg="red4").place(x=30,y=160)
lbl_email=Entry(rightFrame,font=("times new roman",15),bg="white")
lbl_email.place(x=200,y=160,width=400,height=35)
        

lbl_id=Label(rightFrame,text="Valid ID No :",font=("times new roman",20,"bold"),fg="white",bg="red4").place(x=30,y=210)
lbl_id=Entry(rightFrame,font=("times now roman",15),bg="white",textvariable=e_entryno)
lbl_id.place(x=200,y=210,width=400,height=35)



#...........................................................excel File.................................................................
#Excel Data Sheet Creation 
wb = Workbook()
ws = wb.active    

def save_inputintoexcel(): 

    # Get the text input from user and save it in a list.  
    user_name = lbl_name.get() 
    user_add = lbl_add.get()
    user_no  = lbl_no.get()
    user_email = lbl_email.get()
    user_foodcost = costoffoodvar.get()
    user_roomcost = costofroomvar.get()
    user_totalbill = totalcostvar.get()
    user_billno = lbl_id.get()
    
    sheet = wb.active
    sheet.cell(row=1 ,column = 1).value ='Name  OF customer'
    sheet.column_dimensions['A'].width=25
    sheet.cell(row=1 ,column = 2).value ='Address'
    sheet.column_dimensions['B'].width=25
    sheet.cell(row=1 ,column =3 ).value ='Contact No'
    sheet.column_dimensions['C'].width=15
    sheet.cell(row=1 ,column = 4).value ='Email ID'
    sheet.column_dimensions['D'].width=25
    sheet.cell(row=1 ,column = 5).value ='Cost Of Food'
    sheet.column_dimensions['E'].width=15
    sheet.cell(row=1 ,column = 6).value ='Cost Of Room'
    sheet.column_dimensions['F'].width=15    
    sheet.cell(row=1 ,column = 7).value ='Total Cost'
    sheet.column_dimensions['G'].width=15
    sheet.cell(row=1 ,column = 8).value ='ID  No'
    sheet.column_dimensions['H'].width=15
    sheet.cell(row=1 ,column = 9).value ='Date'
    sheet.column_dimensions['I'].width=15
    # Append the user input into excel file in a list one by one row wise using openpyxl library.  
    ws.append([user_name ,user_add,user_no,user_email,user_foodcost,user_roomcost,user_totalbill,user_billno,date])

    # Save the excel file with given name and location path.  
    wb.save("user_inputs_list_column.xlsx") 



 # Create a button to call above function to save data into excel file when clicked on it .
button= Button(buttonFrame, text='Save into Excel', font=('arial', 14, 'bold'), fg='white', bg='red4', bd=3, padx=5,
                     command=save_inputintoexcel)
button.grid(row=0, column=4)
  														    

root.mainloop()